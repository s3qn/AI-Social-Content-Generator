from apify_client import ApifyClient
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timezone
import os
import json
import argparse
import re
import requests

CACHE_DIR = Path(__file__).resolve().parents[3] / "cache"
VIRAL_THUMBS_DIR = CACHE_DIR / "viral_thumbs"
VIRAL_VIDEOS_DIR = CACHE_DIR / "viral_videos"
VIRAL_MEDIA_TIMEOUT = 30


VIRAL_EXCEL_HEADERS = [
    "Tier",
    "Post Date",
    "Username",
    "Language",
    "Views",
    "Likes",
    "Comments",
    "Shares",
    "Engagement Score",
    "Caption",
    "Post URL",
]

VIRAL_EXCEL_FIELD_MAP = {
    "Tier": "tier",
    "Post Date": "post_date",
    "Username": "username",
    "Language": "lang",
    "Views": "views",
    "Likes": "likes",
    "Comments": "comments",
    "Shares": "shares",
    "Engagement Score": "engagement_score",
    "Caption": "caption",
    "Post URL": "post_url",
}

VIRAL_EXCEL_COLUMN_WIDTHS = {
    "Tier": 10,
    "Post Date": 12,
    "Username": 22,
    "Language": 10,
    "Views": 12,
    "Likes": 10,
    "Comments": 12,
    "Shares": 10,
    "Engagement Score": 18,
    "Caption": 60,
    "Post URL": 50,
}


VIRAL_ACTOR_ID = "patient_discovery/instagram-search-reels"
VIRAL_PAGES_PER_KEYWORD = 2
VIRAL_TOP_BIGGEST = 3
VIRAL_TOP_RESONANT = 2
# Floor for the resonant tier: a 40-view post with 3 comments has a huge
# ratio but is noise, not resonance (the old ratio-only ranking's failure
# mode). Starting constant — tune after eyeballing a real report.
VIRAL_RESONANT_MIN_VIEWS = 1000


def load_apify_api_key():

    load_dotenv(find_dotenv())
    apify_api_key = os.getenv("APIFY_API_KEY")
    return apify_api_key

def convert_to_json(data, handle, suffix):

    with open(Path(__file__).resolve().parents[3] / "cache" / f"{handle}-{suffix}.json", "w") as f:
        json.dump(data, f, indent=2, default=str)
    return None

def fetch_posts(apify_api_key, handle, limit=20):

    client = ApifyClient(apify_api_key)
    actor_client = client.actor('apify/instagram-scraper')
    result = actor_client.call(run_input={
        "directUrls": [f"https://www.instagram.com/{handle}/"],
        "resultsType": "posts",
        "resultsLimit": limit,
    })

    if result is None:
        print('Actor run failed...')
        return

    data = client.dataset(result["defaultDatasetId"])
    data_list = data.list_items().items
    return data_list

def fetch_profile(apify_api_key, handle):

    client = ApifyClient(apify_api_key)
    actor_client = client.actor('apify/instagram-scraper')
    result = actor_client.call(run_input={
        "directUrls": [f"https://www.instagram.com/{handle}/"],
        "resultsType": "details",
        "resultsLimit": 1,
    })

    if result is None:
        print('Actor run failed...')
        return

    # FOR DEBUG: print(f"Result keys: {result.keys()}")

    data = client.dataset(result["defaultDatasetId"])
    data_list = data.list_items().items
    return data_list

def get_profile(user, limit):

    cache_dir = Path(__file__).resolve().parents[3] / "cache"
    profile_path = cache_dir / f"{user}-profile.json"
    
    if profile_path.exists():
        print(f"Found cache on disk for @{user}, skipping scrape")
        profile_data = json.loads(profile_path.read_text(encoding="utf-8"))
        return profile_data[0]
    
    print("Cache not found... proceeding to scrape")
    # Load API Key

    print("Loading API Key...")
    apify_api_key = load_apify_api_key()

    # Scrape handle from the gram

    print(f"Scraping for @{user}")
    print(f"Scraping {limit} posts for @{user}...")
    post_data = fetch_posts(apify_api_key, user, limit)

    print(f"Scraping profile for @{user}...")
    profile_data = fetch_profile(apify_api_key, user)
    
    print("Saving to cache...")
    convert_to_json(post_data, user, "posts")
    convert_to_json(profile_data, user, "profile")

    print(f"Done! Saved {len(post_data)} posts and profile for @{user}!")
    return profile_data[0]

def _sanitize_keyword_for_filename(keyword: str) -> str:
    """Make a keyword safe for use as a filename. Replace whitespace,
    quotes, slashes, and other unsafe chars with underscores. Hebrew
    and other Unicode letters are kept as-is (Linux fs supports them)."""
    safe = re.sub(r'[\s/\\\'"<>:|?*]+', '_', keyword)
    return safe[:80]


def viral_cache_path(keyword: str) -> Path:
    safe = _sanitize_keyword_for_filename(keyword)
    return Path(__file__).resolve().parents[3] / "cache" / f"viral_{safe}.json"


def fetch_viral_reels(apify_api_key: str, keyword: str) -> list[dict]:
    """Call patient_discovery/instagram-search-reels Actor for one
    keyword. Returns raw post list (no dedup, no filter, no rank)."""
    client = ApifyClient(apify_api_key)
    actor_client = client.actor(VIRAL_ACTOR_ID)
    result = actor_client.call(run_input={
        "query": keyword,
        "maxPages": VIRAL_PAGES_PER_KEYWORD,
    })

    if result is None:
        print(f"Viral actor run failed for keyword={keyword!r}")
        return []

    data = client.dataset(result["defaultDatasetId"])
    return data.list_items().items


def get_viral_reels(keyword: str, force_refresh: bool = False) -> list[dict]:
    """Public entry. Cache-first (mirrors get_profile). Returns raw post
    list. force_refresh=True bypasses cache."""
    cache_path = viral_cache_path(keyword)

    if cache_path.exists() and not force_refresh:
        print(f"Found viral cache on disk for keyword={keyword!r}, skipping scrape")
        return json.loads(cache_path.read_text(encoding="utf-8"))

    print(f"Viral cache not found for keyword={keyword!r}, scraping...")
    apify_api_key = load_apify_api_key()
    raw = fetch_viral_reels(apify_api_key, keyword)

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(raw, f, indent=2, default=str)

    print(f"Saved {len(raw)} viral reels for keyword={keyword!r}")
    return raw


def dedup_viral_posts(posts: list[dict]) -> list[dict]:
    """Dedup by `code`. Keep first occurrence."""
    seen: set[str] = set()
    out: list[dict] = []
    for p in posts:
        code = p.get("code")
        if code and code not in seen:
            seen.add(code)
            out.append(p)
    return out


def viral_views(post: dict) -> int:
    return post.get("ig_play_count") or post.get("play_count") or 0


def compute_viral_engagement_score(post: dict) -> float:
    """(shares + comments) / views. Returns 0 if views missing/zero."""
    views = viral_views(post)
    if views <= 0:
        return 0.0
    shares = post.get("share_count") or 0
    comments = post.get("comment_count") or 0
    return (shares + comments) / views


def tier_and_rank_viral(posts: list[dict]) -> list[dict]:
    """Two tiers per keyword, no overlap:
    🏆 biggest:  top VIRAL_TOP_BIGGEST by raw views (what blew up).
    💬 resonant: from the REMAINING posts, top VIRAL_TOP_RESONANT by
       engagement ratio, with a minimum-views floor (what resonates).
    Ranking is purely by these metrics — language is display-only
    downstream, never a sort key or filter (ideas translate; best
    content wins regardless of market). Tags each post with
    `_viral_tier` = "biggest"/"resonant"."""
    if not posts:
        return []

    by_views = sorted(posts, key=viral_views, reverse=True)
    biggest = by_views[:VIRAL_TOP_BIGGEST]
    biggest_codes = {p.get("code") for p in biggest}

    candidates = [
        p for p in posts
        if p.get("code") not in biggest_codes
        and viral_views(p) >= VIRAL_RESONANT_MIN_VIEWS
    ]
    resonant = sorted(
        candidates, key=compute_viral_engagement_score, reverse=True
    )[:VIRAL_TOP_RESONANT]

    for p in biggest:
        p["_viral_tier"] = "biggest"
    for p in resonant:
        p["_viral_tier"] = "resonant"

    return biggest + resonant


def viral_post_key(post: dict) -> str:
    """Stable filename key for a post's downloaded media. `pk` is None
    in real actor output; `id` is the reliable field, `code` the last
    resort."""
    return str(post.get("pk") or post.get("id") or post.get("code") or "")


def viral_thumb_path(post_key: str) -> Path:
    return VIRAL_THUMBS_DIR / f"{post_key}.jpg"


def viral_video_path(post_key: str) -> Path:
    return VIRAL_VIDEOS_DIR / f"{post_key}.mp4"


def _download_file(url: str, dest: Path) -> bool:
    """Best-effort download. Returns True on success, never raises."""
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        resp = requests.get(url, timeout=VIRAL_MEDIA_TIMEOUT)
        resp.raise_for_status()
        dest.write_bytes(resp.content)
        return True
    except Exception as e:
        print(f"viral media download failed ({dest.name}): {e}")
        if dest.exists():
            dest.unlink()
        return False


def download_viral_media(top_posts: list[dict]) -> None:
    """Fetch media for the tiered posts AT SCRAPE TIME, while the CDN
    URLs are minutes old (they 403 within hours — an on-demand download
    later is impossible). Thumbnails for ALL top posts; videos only for
    Biggest-tier posts with original audio (licensed music = lyrics
    junk, not worth transcribing). Per-file failures log and skip —
    never fail the pipeline. Skips files already on disk (cache hits)."""
    for post in top_posts:
        key = viral_post_key(post)
        if not key:
            continue

        thumb_url = post.get("thumbnail_url")
        thumb_dest = viral_thumb_path(key)
        if thumb_url and not thumb_dest.exists():
            _download_file(thumb_url, thumb_dest)

        audio_type = (post.get("clips_metadata") or {}).get("audio_type") or ""
        if post.get("_viral_tier") == "biggest" and audio_type == "original_sounds":
            video_url = post.get("video_url")
            video_dest = viral_video_path(key)
            if video_url and not video_dest.exists():
                _download_file(video_url, video_dest)


def update_cached_transcript(keyword: str, post_key: str, transcript: str) -> None:
    """Write a lazily-made transcript back into the keyword's cached
    viral JSON (atomic replace) so future reports reuse it instead of
    re-running Whisper."""
    cache_path = viral_cache_path(keyword)
    if not cache_path.exists():
        return
    try:
        posts = json.loads(cache_path.read_text(encoding="utf-8"))
        changed = False
        for p in posts:
            if viral_post_key(p) == post_key:
                p["transcript"] = transcript
                changed = True
        if not changed:
            return
        tmp = cache_path.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(posts, indent=2, default=str), encoding="utf-8")
        tmp.replace(cache_path)
    except Exception as e:
        print(f"transcript write-back failed for keyword={keyword!r}: {e}")


def extract_viral_summary(post: dict, keyword: str) -> dict:
    """Pull flat fields for display + future Excel export."""
    caption_text = ""
    caption_obj = post.get("caption")
    if isinstance(caption_obj, dict):
        caption_text = caption_obj.get("text", "") or ""
    elif isinstance(caption_obj, str):
        caption_text = caption_obj

    user_obj = post.get("user") or {}
    username = user_obj.get("username", "unknown")

    code = post.get("code", "")
    post_url = f"https://www.instagram.com/reel/{code}/" if code else ""

    taken_at = post.get("taken_at")
    if taken_at:
        try:
            post_date = datetime.fromtimestamp(taken_at, tz=timezone.utc).strftime("%Y-%m-%d")
        except (TypeError, ValueError, OSError):
            post_date = "unknown"
    else:
        post_date = "unknown"

    post_key = viral_post_key(post)
    thumb = viral_thumb_path(post_key)
    video = viral_video_path(post_key)

    return {
        "caption": caption_text,
        "likes": post.get("like_count") or 0,
        "comments": post.get("comment_count") or 0,
        "views": post.get("ig_play_count") or post.get("play_count") or 0,
        "shares": post.get("share_count") or 0,
        "username": username,
        "post_url": post_url,
        "post_date": post_date,
        "keyword_source": keyword,
        "engagement_score": compute_viral_engagement_score(post),
        "tier": post.get("_viral_tier", "biggest"),
        # Display-only context — never used for ranking or filtering.
        "lang": post.get("original_lang_for_translations") or "",
        "thumbnail_url": post.get("thumbnail_url") or "",
        # Whisper-evidence rider: logged at import time so the parked
        # video-analysis decision has data on this niche's audio mix.
        "audio_type": (post.get("clips_metadata") or {}).get("audio_type") or "",
        "pk": post_key,
        # Local media downloaded at scrape time (CDN URLs die in hours).
        # Cards must read these, never the CDN URLs.
        "local_thumb": str(thumb) if thumb.exists() else "",
        "local_video": str(video) if video.exists() else "",
        # Filled lazily on the first 💡/🎤 tap, then cached in the JSON.
        # May be a plain string (old) or {"text","segments"} (new).
        "transcript": post.get("transcript") or "",
        # Reel length (seconds) for Phase 3 pacing analysis.
        "video_duration": post.get("video_duration") or 0,
    }


def scrape_and_process_viral_keywords(
    keywords: list[str],
    force_refresh: bool = False,
) -> list[dict]:
    """Full pipeline for multiple keywords. Returns flat list of summary
    dicts (up to 5 per keyword)."""
    all_results: list[dict] = []
    for kw in keywords:
        raw = get_viral_reels(kw, force_refresh=force_refresh)
        deduped = dedup_viral_posts(raw)
        tiered = tier_and_rank_viral(deduped)
        download_viral_media(tiered)
        for post in tiered:
            all_results.append(extract_viral_summary(post, kw))
        print(f"keyword={kw!r}: {len(raw)} raw → {len(deduped)} deduped → {len(tiered)} kept")
    return all_results


def _delete_media_for_cache_file(path: Path) -> None:
    """Remove the downloaded thumbs/videos belonging to one cached viral
    JSON (walks its post keys before the JSON itself is deleted). Keeps
    cache/viral_videos from growing forever."""
    try:
        posts = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return
    for p in posts:
        key = viral_post_key(p)
        if not key:
            continue
        for media in (viral_thumb_path(key), viral_video_path(key)):
            if media.exists():
                media.unlink()


def invalidate_viral_cache(keyword: str | None = None) -> int:
    """Delete cached viral results AND their downloaded media. If
    keyword is None, delete all viral_*.json. Returns count of cache
    files deleted."""
    cache_dir = Path(__file__).resolve().parents[3] / "cache"
    if not cache_dir.exists():
        return 0

    if keyword is not None:
        path = viral_cache_path(keyword)
        if path.exists():
            _delete_media_for_cache_file(path)
            path.unlink()
            return 1
        return 0

    count = 0
    for p in cache_dir.glob("viral_*.json"):
        _delete_media_for_cache_file(p)
        p.unlink()
        count += 1
    return count


def _sanitize_sheet_name(keyword: str) -> str:
    """Excel sheet names: max 31 chars, no /\\?*[]:'. Strip + replace
    invalid chars with underscore. Hebrew/Unicode OK."""
    safe = re.sub(r"[/\\?*\[\]:']+", "_", keyword.strip())
    return safe[:31] if safe else "Sheet"


def viral_excel_path(user_id: int | str) -> Path:
    """Standard location for the viral Excel report. Overwritten each
    generation. Keyed by user_id, NOT handle: two users can manage the
    same IG account (same collision class as the bg/logo fix)."""
    return Path(__file__).resolve().parents[3] / "cache" / f"viral_report_{user_id}.xlsx"


def build_viral_excel(results: list[dict], output_path: Path) -> Path:
    """Build .xlsx file. One sheet per keyword. Returns the output path.

    Within each sheet: 'biggest' tier first (views descending), then
    'resonant' (engagement_score descending). Header row is frozen. If
    results is empty, a single 'No Data' sheet is created so the user
    gets a file back either way.
    """
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    if not results:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No viral posts found for any keyword."
        ws["A2"] = "Try different keywords or check your network."
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path

    grouped: dict[str, list[dict]] = {}
    for r in results:
        kw = r.get("keyword_source", "unknown")
        grouped.setdefault(kw, []).append(r)

    used_names: set[str] = set()

    for keyword, rows in grouped.items():
        base_name = _sanitize_sheet_name(keyword)
        sheet_name = base_name
        suffix = 2
        while sheet_name in used_names:
            sheet_name = f"{base_name[:28]}_{suffix}"
            suffix += 1
        used_names.add(sheet_name)

        ws = wb.create_sheet(sheet_name)

        for col_idx, header in enumerate(VIRAL_EXCEL_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        sorted_rows = sorted(
            rows,
            key=lambda r: (
                (0, -(r.get("views") or 0))
                if r.get("tier") == "biggest"
                else (1, -(r.get("engagement_score") or 0))
            ),
        )

        for row_idx, row_data in enumerate(sorted_rows, start=2):
            for col_idx, header in enumerate(VIRAL_EXCEL_HEADERS, start=1):
                key = VIRAL_EXCEL_FIELD_MAP[header]
                value = row_data.get(key, "")
                if key == "engagement_score" and isinstance(value, (int, float)):
                    value = f"{value:.4f}"
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, header in enumerate(VIRAL_EXCEL_HEADERS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = VIRAL_EXCEL_COLUMN_WIDTHS.get(header, 15)

        ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():

    parser = argparse.ArgumentParser(description="Fetch instagram posts from handle")
    parser.add_argument("--user", help="Instagram Handle to fetch (e.g: nasa)" ,default="inna_cheskis")
    parser.add_argument("-p" ,"--post-number", choices=range(1, 21), type=int, help="How many posts to fetch from profile (1-20)", default=1)
    args = parser.parse_args()
    user = args.user
    limit = args.post_number

    # Scrape Content
    get_profile(user, limit)

if __name__ == "__main__":
    main()